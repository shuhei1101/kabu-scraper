import pandas as pd
from openpyxl import load_workbook

class StockXlsxHandler:
    '''stock_table.xlsx を操作するクラス'''

    def __init__(self, xlsx_path, key):
        '''
        :raise ValueError: 指定したキーが存在しない場合
        '''
        self.xlsx_path = xlsx_path
        self.key = key
        self.df = pd.read_excel(xlsx_path, engine='openpyxl')

        # キーの存在確認
        if self.key not in self.df.columns:
            raise ValueError(f"キー '{self.key}' は存在しません。")

    def get_record(self, key: str):
        '''id（コード）を指定してレコード取得'''
        match = self.df[self.df[self.key].astype(str) == str(key)]
        if match.empty:
            raise ValueError(f"コード '{key}' のレコードが見つかりません。")
        record = match.iloc[0]
        record = record.where(pd.notnull(record), None)
        return record

    def get_keys(self):
        '''キーのデータを取得'''
        return self.df[self.key].tolist()

    def get_column_data(self, column_name: str):
        '''指定したカラムのデータを取得'''
        if column_name not in self.df.columns:
            raise ValueError(f"カラム '{column_name}' は存在しません。")
        return self.df[column_name].tolist()

    def update_record(self, updated_record: pd.Series):
        '''レコードを渡して一行更新（書式維持）'''
        # Excelファイルをロード
        wb = load_workbook(self.xlsx_path, data_only=False)
        ws = wb.active

        # ヘッダー読み取り
        headers = [cell.value for cell in ws[1]]
        if self.key not in headers:
            raise ValueError(f"キー '{self.key}' はシートに存在しません。")

        key_col_idx = headers.index(self.key) + 1  # Excel列は1始まり

        # 該当行を探して更新
        for row in ws.iter_rows(min_row=2):
            cell = row[key_col_idx - 1]
            if cell.value == updated_record[self.key]:
                for idx, header in enumerate(headers):
                    if header in updated_record:
                        target_cell = row[idx]
                        # 元々関数なら上書きしない
                        if not (target_cell.data_type == 'f'):  # 'f'はformulaの略
                            target_cell.value = updated_record[header]
                break
        else:
            raise ValueError(f"コード '{updated_record[self.key]}' のレコードが存在しません。")

        # 書式を維持したまま保存
        wb.save(self.xlsx_path)
        wb.close()

# 動作確認用
if __name__ == "__main__":
    handler = StockXlsxHandler(
        xlsx_path='/Users/nishikawashuhei/Downloads/配当管理.xlsx',
        key='No'
    )
    try:
        record = handler.get_record('1.0')
        print("取得したレコード:", record)

        stock_codes = handler.get_keys()
        print("銘柄コードのリスト:", stock_codes)

        # 更新するレコードの例
        # record['銘柄'] = 'testest'
        # handler.update_record(record)
    except ValueError as e:
        print(e)
